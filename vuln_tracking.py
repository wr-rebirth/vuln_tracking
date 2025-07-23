import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment
from openpyxl.utils import get_column_letter
import sys

if len(sys.argv) < 3:
    print("用法: python3 vuln_tracking.py <输入文件路径> <输出文件路径>")
    sys.exit(1)

srcfile = sys.argv[1]
output_file = sys.argv[2]

# 读取原始表格
df = pd.read_excel(srcfile, dtype=str, engine="openpyxl")
df.fillna("", inplace=True)
# 去除所有字段的前后空格
df = df.applymap(lambda x: x.strip() if isinstance(x, str) else x)

# 将漏洞等级列中的紧急都替换为严重
df['漏洞等级'] = df['漏洞等级'].replace('紧急', '严重')

# 清理漏洞位置：去掉 http/https 开头的结尾 "/"
df["漏洞位置"] = df["漏洞位置"].apply(
    lambda x: x[:-1] if x.startswith("http") and x.endswith("/") else x
)

# 获取每个系统下的所有测试时间 + 最后测试时间
system_test_times = (
    df.groupby("系统")["测试时间"]
    .apply(lambda x: sorted(set(x)))
    .to_dict()
)

system_last_test_time = {
    system: times[-1] for system, times in system_test_times.items()
}

# 添加组合键：系统 + 地址 + 漏洞名称
df["组合键"] = df["系统"] + "|" + df["漏洞位置"] + "|" + df["漏洞名称"]

results = []

# 分组处理漏洞项
for key, group in df.groupby("组合键"):
    group = group.sort_values("测试时间")
    system = group.iloc[0]["系统"]
    address = group.iloc[0]["漏洞位置"]
    vuln = group.iloc[0]["漏洞名称"]
    level = group.iloc[0]["漏洞等级"]
    vuln_test_times = sorted(set(group["测试时间"]))  # 漏洞出现的时间
    all_system_times = system_test_times[system]     # 该系统所有测试时间
    system_last_time = system_last_test_time[system] # 该系统的最后测试时间

    discovery_time = vuln_test_times[0]
    last_seen_time = vuln_test_times[-1]
    retest_time_str = "\n".join(all_system_times)

    # 判断该漏洞是否出现在系统的最后一次测试中，并计算复测通过时间
    # 找到最后一次出现时间在系统所有测试时间中的下一个时间点
    passed_time = ""
    if last_seen_time in all_system_times:
        idx = all_system_times.index(last_seen_time)
        if idx + 1 < len(all_system_times):
            next_time = all_system_times[idx + 1]
            # 判断该漏洞在下一个时间点是否还出现
            if next_time not in vuln_test_times:
                passed_time = next_time
    
    if system_last_time in vuln_test_times:
        fix_status = "未修复"
    else:
        fix_status = "已修复"

    results.append({
        "系统": system,
        "漏洞位置": address,
        "漏洞名称": vuln,
        "漏洞等级": level,
        # "测试时间": last_seen_time,
        "漏洞修复情况": fix_status,
        "发现时间": discovery_time,
        "复测时间点": retest_time_str,
        "复测通过时间": passed_time,
        # "最后测试时间": system_last_time,
        "漏洞描述": group.iloc[0]["漏洞描述"],
        "漏洞危害": group.iloc[0]["漏洞危害"],
        "修复方案": group.iloc[0]["修复方案"],
    })

# 保存处理结果
result_df = pd.DataFrame(results)
# 保存结果
result_df.to_excel(output_file, index=False)

# 加载 Excel 文件
wb = load_workbook(output_file)
ws = wb.active

# 定义颜色
color_dict = {
    '严重': PatternFill(fill_type='solid', fgColor='FFA500'),   # 橙色
    '高危': PatternFill(fill_type='solid', fgColor='FFA500'),   # 橙色
    '中危': PatternFill(fill_type='solid', fgColor='FFD580'),   # 浅橙色
    '低危': PatternFill(fill_type='solid', fgColor='D3D3D3'),   # 浅灰色
    '未修复': PatternFill(fill_type='solid', fgColor='FFA500'), # 橙色
    '已修复': PatternFill(fill_type='solid', fgColor='90EE90'), # 浅绿色
}

# 自动查找列号
header = [cell.value for cell in ws[1]]
level_col = header.index('漏洞等级')
status_col = header.index('漏洞修复情况')

for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    # 漏洞等级
    level = row[level_col].value
    if level in color_dict:
        row[level_col].fill = color_dict[level]
    # 漏洞修复情况
    status = row[status_col].value
    if status in color_dict:
        row[status_col].fill = color_dict[status]

# 设置自动换行和固定行高
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    ws.row_dimensions[row[0].row].height = 16  # 固定行高
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)  # 自动换行

# 指定每一列的宽度
col_widths = {
    '系统': 20,
    '漏洞位置': 30,
    '漏洞名称': 70,
    '漏洞等级': 10,
    '漏洞修复情况': 15,
    '发现时间': 15,
    '复测时间点': 15,
    '复测通过时间': 15
}

# 获取表头
header = [cell.value for cell in ws[1]]

for idx, col_name in enumerate(header, 1):
    if col_name in col_widths:
        ws.column_dimensions[get_column_letter(idx)].width = col_widths[col_name]

# 前三列靠左，后五列居中
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    for idx, cell in enumerate(row, 1):
        if idx <= 3:
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        else:
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

wb.save(output_file)